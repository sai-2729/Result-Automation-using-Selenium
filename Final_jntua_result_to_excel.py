from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import time

from tkinter import *

jntua = webdriver.Chrome(executable_path="webdrivers/chromedriver.exe")


def fetch():
    xl = openpyxl.Workbook()

    sheet = xl.active

    # ----- year -------

    b_year = year.get()
    # -----------------

    # ------ sem -------

    b_sem = sem.get()

    # -----------------

    # ----- branch -----

    b_branch = branch.get()
    # -----------------

    # ------------ Subjects -----------------

    if b_branch == "CSE":
        if b_year == 1 and b_sem == 1:
            cse = ['CP LAB', 'ELCS LAB', 'ED', 'PHYSICS', 'CP', 'M-1', 'ENGLISH']
        elif b_year == 1 and b_sem == 2:
            cse = ['ENG & IT WORKSHOP', 'CHEMISTRY LAB', 'DS LAB', 'ES', 'CHEMISTRY', 'DS', 'M-2', 'ENGLISH']
        elif b_year == 2 and b_sem == 1:
            cse = ['BEEE LAB', 'BEEE', 'M-3', 'MEFA', 'DBMS LAB', 'DM', 'DBMS', 'DLD']
        elif b_year == 2 and b_sem == 2:
            cse = ['P AND S', 'ONLINE TEST 1', 'JAVA LAB', 'FLAT', 'OOPJ', 'CO', 'SE', 'MPI LAB', 'MPI']
        elif b_year == 3 and b_sem == 1:
            cse = ['SOCIAL VALUES', 'OS LAB', 'OOAD LAB', 'BD', 'ST', 'PPL', 'OOAD', 'CN', 'OS']
        elif b_year == 3 and b_sem == 2:
            cse = ['ENGLISH LAB', 'ONLINE TEST 2', 'DWDM LAB', 'WIT LAB', 'WIT', 'DAA', 'DP', 'DWDM', 'CD',
                   'IPR']  # Subject Names
        elif b_year == 4 and b_sem == 1:
            cse = ['MS', 'MAD LAB', 'GCC LAB', 'RTS', 'SA', 'MAD', 'IS', 'GCC']
        else:
            cse = ['PROJECT WORK', 'TECHNICAL SEMINAR', 'VIVA VOCE', 'CS', 'MC']

        subject=cse.copy()

    # ---------------------------------------



    link = url_link.get()

    rline = 3

    start_hall = hall_ticket = start_hallticket.get()
    end_hall = end_hallticket.get()
    start_limit = int(start_hall[8:])
    end_limit = int(end_hall[8:])
    limit = end_limit - start_limit + 1
    code = int(hall_ticket[7:])

    rows=len(subject)
    column=subject*3
    
    for i in range(1, rows * 3, 3):
        sheet.cell(row=2, column=i + 1).value = 'MID'
        sheet.cell(row=2, column=i + 2).value = 'SEM'  # sub heading of subjects
        sheet.cell(row=2, column=i + 3).value = 'P/F'

    for i in range(1, rows * 3, 3):
        sheet.cell(row=1, column=i + 2).value = subject.pop()  # Subjects names Entering into Excel file
    
    for i in range(limit):

        hall_ticket = hall_ticket[:7] + str(code)

        jntua.get(link)
        jntua.find_element_by_class_name('txt').send_keys(hall_ticket)
        jntua.find_element_by_class_name('ci').click()

        code += 1
        try:
            jntua.implicitly_wait(10)
            jntua.find_element_by_xpath('/html/body/div/div[1]/div/div/center/div[1]/table')

        except NoSuchElementException:
            continue

        sheet.cell(row=rline, column=1).value = hall_ticket

        rows = len(jntua.find_elements_by_xpath(
            '//*[@id="rs"]/table/tbody/tr'))  # len of rows in tables also known as Subject length if "rows-2"
        cols = len(jntua.find_elements_by_xpath('//*[@id="rs"]/table/tbody/tr[1]/th'))  # len of the columns

        cline = 2

        for r in range(2, rows):
            for c in range(3, 7):
                if c == 5:
                    continue
                data = jntua.find_element_by_xpath(
                    "//*[@id='rs']/table/tbody/tr[" + str(r) + "]/td[" + str(c) + "]").text
                sheet.cell(row=rline, column=cline).value = data  # storing Marks into Excel file
                cline = cline + 1

        rline = rline + 1



    txt.delete(0.0, 'end')
    txt.insert(0.0, "\nData Entered Successfully....\n")
    txt.insert(0.0, '============ Log =============\n')

    file = "results/" + str(b_year)
    xl.save(file + " Year " + str(b_sem) + " Sem " + b_branch + " Results.xlsx")


def jntuaresults():
    jntua.get("https://jntuaresults.ac.in/")


gui = Tk()

gui.geometry('1000x500+430+250')

gui.title("SVIT JNTUA SEMESTER RESULTS AUTOMATION TOOL")
Label(text="Sri Venkateswara Institute Of Technology (SVIT)", fg="black", font=('Comic Sans MS', 15)).place(x=300, y=1)

# -------- Year ------

Label(text="Year --->  ", font=13).place(x=100, y=100)
year = IntVar()
Radiobutton(gui, text="1", variable=year, value=1, font=10).place(x=220, y=100)
Radiobutton(gui, text="2", variable=year, value=2, font=10).place(x=300, y=100)
Radiobutton(gui, text="3", variable=year, value=3, font=10).place(x=380, y=100)
Radiobutton(gui, text="4", variable=year, value=4, font=10).place(x=460, y=100)
# --------------------

# -------- Sem -------

Label(text="Sem --->  ", font=13).place(x=100, y=150)
sem = IntVar()
Radiobutton(gui, text="1", variable=sem, value=1, font=10).place(x=220, y=150)
Radiobutton(gui, text="2", variable=sem, value=2, font=10).place(x=300, y=150)

# -------------------


# ----------- Branch -------------

branch = StringVar()
branch.set(" ")
Label(text="Branch --->  ", font=13).place(x=100, y=200)
Radiobutton(gui, text="CSE", variable=branch, value="CSE", font=10).place(x=220, y=200)
Radiobutton(gui, text="ECE", variable=branch, value="ECE", font=10).place(x=300, y=200)
Radiobutton(gui, text="EEE", variable=branch, value="EEE", font=10).place(x=380, y=200)
Radiobutton(gui, text="Civil", variable=branch, value="Civil", font=10).place(x=460, y=200)
Radiobutton(gui, text="Mech", variable=branch, value="Mech", font=10).place(x=540, y=200)

# --------------------------------

# -------- Hallticket no. ---------

start_hallticket = StringVar()
end_hallticket = StringVar()

Label(text="Hallticket No. --->", font=13).place(x=100, y=250)

Label(text="Start : ", font=10).place(x=300, y=250)
Entry(gui, justify="center", textvariable=start_hallticket).place(x=380, y=255)

Label(text="End : ", font=10).place(x=520, y=250)
Entry(gui, justify="center", textvariable=end_hallticket).place(x=600, y=255)
# ---------------------------------

# ------- External Entry ---------

ext_entry = StringVar()
Label(text="External Entry (optional) --->", font=13).place(x=100, y=300)
Entry(gui, justify="center", textvariable=ext_entry).place(x=380, y=305)

Button(gui, text="ADD", command=fetch).place(x=550, y=300)
# ---------------------------------

# ---------- Get URl -------------

url_link = StringVar()

Label(text="URL --->", font=13).place(x=100, y=350)
Entry(gui, justify="center", width=50, textvariable=url_link).place(x=200, y=355)

Button(gui, text="Get Link", command=jntuaresults).place(x=550, y=350)

# -------------------------------

# -------- Fetch Button ---------

Button(gui, text="Fetch It", command=fetch).place(x=400, y=400)
# -------------------------------
# ---------- Log ----------------

txt = Text(gui, width=30, height=5)
txt.place(x=740, y=60)
txt.insert(0.0, '============ Log =============')

# -------------------------------

mainloop()
# -------------------------------------------------------------------------------------------------------------------
