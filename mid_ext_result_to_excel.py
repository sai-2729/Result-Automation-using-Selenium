from selenium import webdriver
import openpyxl



xl=openpyxl.Workbook()
sh=xl.active
sh.title("Sheet")



for i in range(1,27,3):
    c=sh.cell(2,i+1)
    c.value="MID"
    c=sh.cell(2,i+2)
    c.value="EXTERNAL"
    c=sh.cell(2,i+3)
    c.value="RESULT"
    
    

xl.save("xcel.xlsx")
