from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import time

path="F:\marks1.xlsx"                                             #Getting path of Excel file in system
workbook=openpyxl.load_workbook(path)
sheet=workbook.active

browser = webdriver.Chrome(executable_path='C:\webdrivers\chromedriver.exe')        #Activating Chrome
arr=['OS','CN','OOAD','PPL','ST','BD','OOAD LAB',' OS LAB','SOCIAL VALUES']
arr.reverse()                                                                     #Subject Names

rline=3
hallticketno="179F1A0500"
for i in range(0,45):
    num=hallticketno[7:]
    num=int(num)+1
    hallticketno=hallticketno[0:len(hallticketno)-3]+str(num)                   #Incrementing the Hallticket nos
    print(hallticketno)
    sheet.cell(row=rline,column=1).value=hallticketno
    
    browser.get('https://jntuaresults.ac.in/view-results-56736237.html')          #Getting into results page

    search=browser.find_element_by_xpath('/html/body/div/div[1]/div/div/center/table/tbody/tr/th/center/input[2]')  #Finding Search Button

    hallticket=browser.find_element_by_id('ht')                         #Finding hallticket entering field in website
    hallticket.send_keys(hallticketno)                           # Entering the Hall ticket Automatically
    search.click()

    try:                                             #check if table exists for a Hallticket no
        browser.implicitly_wait(10)
        browser.find_element_by_xpath('/html/body/div/div[1]/div/div/center/div[1]/table')     
    except NoSuchElementException:                  #Exception for neglecting the error
        print('Student Details Not Found')
        continue

    rows=len(browser.find_elements_by_xpath('//*[@id="rs"]/table/tbody/tr'))          #len of rows in tables also known as Subject length if "rows-2"
    cols=len(browser.find_elements_by_xpath('//*[@id="rs"]/table/tbody/tr[1]/th'))    #len of the columns 
    
    cline=2

    for r in range(2,rows):
        for c in range(3,7):
            if c==5:
                continue
            data=browser.find_element_by_xpath("//*[@id='rs']/table/tbody/tr["+str(r)+"]/td["+str(c)+"]").text
            sheet.cell(row=rline,column=cline).value=data                           #storing Marks into Excel file 
            print(data,end=' ')
            cline=cline+1
        print('')

    rline=rline+1

    time.sleep(1)
for i in range(1,(rows-2)*3,3):
    sheet.cell(row=2,column=i+1).value='MID'            
    sheet.cell(row=2,column=i+2).value='SEM'                     #sub heading of subjects
    sheet.cell(row=2,column=i+3).value='P/F'

for i in range(1,(rows-2)*3,3):
    sheet.cell(row=1,column=i+2).value=arr.pop()                  #Subjects names Entering into Excel file

workbook.save(path)